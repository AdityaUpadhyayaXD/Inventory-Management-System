import math
import sys
from typing import Optional, List, Tuple
from PySide6.QtCore import Qt, QSize, QThread, Signal, QTimer

from PySide6.QtWidgets import (
	QMainWindow,
	QWidget,
	QVBoxLayout,
	QHBoxLayout,
	QLabel,
	QPushButton,
	QTableWidget,
	QTableWidgetItem,
	QMessageBox,
	QLineEdit,
	QComboBox,
	QDialog,
	QFormLayout,
	QDoubleSpinBox,
	QSpinBox,
	QStackedWidget,
	QFrame,
	QHeaderView,
	QAbstractItemView,
	QFileDialog,
	QStyledItemDelegate,
)
from PySide6.QtGui import QFont, QColor
from mysql.connector import Error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import AuthenticationDatabase for typing AND the shared LoadingScreen widget
from authentication import AuthenticationDatabase, LoadingScreen, FocusGlowFilter, apply_focus_glow


# ---------------------------------------------------------------------------
# Background workers for paginated data loading
# ---------------------------------------------------------------------------

class _InventoryLoadWorker(QThread):
	"""Fetches one page of inventory products off the main thread."""

	finished = Signal(list, int, int)   # rows, current_page, total_pages
	error    = Signal(str)

	def __init__(
		self,
		database: AuthenticationDatabase,
		page: int,
		page_limit: int,
		search_text: str,
		selected_category: str,
		sort_idx: int,
	) -> None:
		super().__init__()
		self._db            = database
		self._page          = page
		self._page_limit    = page_limit
		self._search_text   = search_text
		self._category      = selected_category
		self._sort_idx      = sort_idx

	def run(self) -> None:
		try:
			connection = self._db._connect()
			cursor = connection.cursor()

			# Count
			count_query = "SELECT COUNT(*) FROM products WHERE 1=1"
			count_params: list = []
			if self._search_text:
				count_query += " AND (sku LIKE %s OR name LIKE %s)"
				count_params.extend([f"%{self._search_text}%", f"%{self._search_text}%"])
			if self._category and self._category != "All Categories":
				if self._category == "Unassigned":
					count_query += " AND (category IS NULL OR category = '')"
				else:
					count_query += " AND category = %s"
					count_params.append(self._category)
			cursor.execute(self._db._format_query(count_query), tuple(count_params))
			total_records = cursor.fetchone()[0]
			total_pages = max(1, math.ceil(total_records / self._page_limit))
			page = max(1, min(self._page, total_pages))

			# Data
			query = "SELECT id, sku, name, category, quantity, unit_price, min_stock, supplier FROM products WHERE 1=1"
			params: list = []
			if self._search_text:
				query += " AND (sku LIKE %s OR name LIKE %s)"
				params.extend([f"%{self._search_text}%", f"%{self._search_text}%"])
			if self._category and self._category != "All Categories":
				if self._category == "Unassigned":
					query += " AND (category IS NULL OR category = '')"
				else:
					query += " AND category = %s"
					params.append(self._category)
			sort_mapping = {
				0: "sku ASC",   1: "sku DESC",
				2: "name ASC",  3: "name DESC",
				4: "quantity ASC", 5: "quantity DESC",
				6: "unit_price ASC", 7: "unit_price DESC",
			}
			query += f" ORDER BY {sort_mapping.get(self._sort_idx, 'sku ASC')}"
			offset = (page - 1) * self._page_limit
			query += " LIMIT %s OFFSET %s"
			params.extend([self._page_limit, offset])
			cursor.execute(self._db._format_query(query), tuple(params))
			rows = cursor.fetchall()
			cursor.close()
			connection.close()
			self.finished.emit(rows, page, total_pages)
		except Exception as exc:
			self.error.emit(str(exc))


class _UsersLoadWorker(QThread):
	"""Fetches one page of registered users off the main thread."""

	finished = Signal(list, int, int)   # rows, current_page, total_pages
	error    = Signal(str)

	def __init__(self, database: AuthenticationDatabase, page: int, page_limit: int, search_text: str = "") -> None:
		super().__init__()
		self._db         = database
		self._page       = page
		self._page_limit = page_limit
		self._search     = search_text

	def run(self) -> None:
		try:
			connection = self._db._connect()
			cursor = connection.cursor()

			# Count
			count_query = "SELECT COUNT(*) FROM users WHERE 1=1"
			count_params = []
			if self._search:
				count_query += " AND (username LIKE %s OR full_name LIKE %s)"
				count_params.extend([f"%{self._search}%", f"%{self._search}%"])

			cursor.execute(self._db._format_query(count_query), tuple(count_params))
			total_records = cursor.fetchone()[0]
			total_pages = max(1, math.ceil(total_records / self._page_limit))
			page = max(1, min(self._page, total_pages))
			offset = (page - 1) * self._page_limit

			# Fetch fields including role
			query = "SELECT id, username, full_name, mobile_number, role, created_at FROM users WHERE 1=1"
			params = []
			if self._search:
				query += " AND (username LIKE %s OR full_name LIKE %s)"
				params.extend([f"%{self._search}%", f"%{self._search}%"])

			query += " ORDER BY created_at ASC LIMIT %s OFFSET %s"
			params.extend([self._page_limit, offset])

			cursor.execute(self._db._format_query(query), tuple(params))
			rows = cursor.fetchall()
			cursor.close()
			connection.close()
			self.finished.emit(rows, page, total_pages)
		except Exception as exc:
			self.error.emit(str(exc))


class AddUserDialog(QDialog):
	def __init__(self, parent: Optional[QWidget] = None) -> None:
		super().__init__(parent)
		self.setWindowTitle("Add New User")
		self.setMinimumSize(400, 360)
		self._build_ui()

	def _build_ui(self) -> None:
		layout = QVBoxLayout(self)
		layout.setContentsMargins(24, 24, 24, 24)
		layout.setSpacing(16)

		title = QLabel("User Account Information")
		title.setStyleSheet("font-size: 18px; font-weight: 700; color: #ffffff;")
		layout.addWidget(title)

		form = QFormLayout()
		form.setVerticalSpacing(12)
		form.setHorizontalSpacing(14)

		self.username_input = QLineEdit()
		self.username_input.setPlaceholderText("Unique username")

		self.fullname_input = QLineEdit()
		self.fullname_input.setPlaceholderText("Full name")

		self.mobile_input = QLineEdit()
		self.mobile_input.setPlaceholderText("10-digit mobile number")

		self.password_input = QLineEdit()
		self.password_input.setPlaceholderText("Minimum 6 characters")
		self.password_input.setEchoMode(QLineEdit.Password)

		self.role_input = QComboBox()
		self.role_input.addItems(["user", "admin"])

		form.addRow("Username:", self.username_input)
		form.addRow("Full Name:", self.fullname_input)
		form.addRow("Mobile No:", self.mobile_input)
		form.addRow("Password:", self.password_input)
		form.addRow("User Role:", self.role_input)
		layout.addLayout(form)

		button_row = QHBoxLayout()
		self.cancel_button = QPushButton("Cancel")
		self.cancel_button.setObjectName("dialogCancelBtn")
		self.cancel_button.clicked.connect(self.reject)

		self.save_button = QPushButton("Add User")
		self.save_button.setObjectName("dialogSaveBtn")
		self.save_button.clicked.connect(self.handle_save)

		button_row.addStretch()
		button_row.addWidget(self.cancel_button)
		button_row.addWidget(self.save_button)
		layout.addLayout(button_row)

		self.setStyleSheet(
			"""
			QDialog {
				background-color: #111827;
			}
			QLabel {
				color: #cbd5e1;
				font-size: 13px;
				font-weight: 500;
			}
			QLineEdit, QComboBox {
				background-color: #1e293b;
				color: #f8fafc;
				border: 1px solid #334155;
				border-radius: 6px;
				padding: 6px 10px;
				font-size: 13px;
				min-height: 20px;
				outline: none;
			}
			QLineEdit:focus, QComboBox:focus {
				border: 1px solid #3b82f6;
			}
			QPushButton {
				border: none;
				border-radius: 6px;
				padding: 10px 18px;
				font-weight: 600;
				font-size: 13px;
				outline: none;
			}
			#dialogSaveBtn {
				background-color: #3b82f6;
				color: white;
			}
			#dialogSaveBtn:hover {
				background-color: #2563eb;
			}
			#dialogCancelBtn {
				background-color: #1e293b;
				color: #cbd5e1;
				border: 1px solid #334155;
			}
			#dialogCancelBtn:hover {
				background-color: #334155;
				color: white;
			}
			"""
		)

	def handle_save(self) -> None:
		username = self.username_input.text().strip().lower()
		fullname = self.fullname_input.text().strip()
		mobile = self.mobile_input.text().strip()
		password = self.password_input.text()
		role = self.role_input.currentText()

		if not username or not fullname or not mobile or not password:
			QMessageBox.warning(self, "Validation Error", "All fields are required to create a user account.")
			return

		if len(password) < 6:
			QMessageBox.warning(self, "Validation Error", "Password must be at least 6 characters.")
			return

		self.save_data = {
			"username": username,
			"fullname": fullname,
			"mobile": mobile,
			"password": password,
			"role": role
		}
		self.accept()


class ResetPasswordDialog(QDialog):
	def __init__(self, username: str, parent: Optional[QWidget] = None) -> None:
		super().__init__(parent)
		self.username = username
		self.setWindowTitle("Reset Password")
		self.setMinimumSize(360, 200)
		self._build_ui()

	def _build_ui(self) -> None:
		layout = QVBoxLayout(self)
		layout.setContentsMargins(24, 24, 24, 24)
		layout.setSpacing(16)

		title = QLabel(f"Reset Password for '{self.username}'")
		title.setStyleSheet("font-size: 16px; font-weight: 700; color: #ffffff;")
		layout.addWidget(title)

		form = QFormLayout()
		self.password_input = QLineEdit()
		self.password_input.setPlaceholderText("Minimum 6 characters")
		self.password_input.setEchoMode(QLineEdit.Password)

		form.addRow("New Password:", self.password_input)
		layout.addLayout(form)

		button_row = QHBoxLayout()
		self.cancel_button = QPushButton("Cancel")
		self.cancel_button.setObjectName("dialogCancelBtn")
		self.cancel_button.clicked.connect(self.reject)

		self.save_button = QPushButton("Reset Password")
		self.save_button.setObjectName("dialogSaveBtn")
		self.save_button.clicked.connect(self.handle_save)

		button_row.addStretch()
		button_row.addWidget(self.cancel_button)
		button_row.addWidget(self.save_button)
		layout.addLayout(button_row)

		self.setStyleSheet(
			"""
			QDialog {
				background-color: #111827;
			}
			QLabel {
				color: #cbd5e1;
				font-size: 13px;
				font-weight: 500;
			}
			QLineEdit {
				background-color: #1e293b;
				color: #f8fafc;
				border: 1px solid #334155;
				border-radius: 6px;
				padding: 6px 10px;
				font-size: 13px;
				min-height: 20px;
				outline: none;
			}
			QLineEdit:focus {
				border: 1px solid #3b82f6;
			}
			QPushButton {
				border: none;
				border-radius: 6px;
				padding: 10px 18px;
				font-weight: 600;
				font-size: 13px;
				outline: none;
			}
			#dialogSaveBtn {
				background-color: #3b82f6;
				color: white;
			}
			#dialogSaveBtn:hover {
				background-color: #2563eb;
			}
			#dialogCancelBtn {
				background-color: #1e293b;
				color: #cbd5e1;
				border: 1px solid #334155;
			}
			#dialogCancelBtn:hover {
				background-color: #334155;
				color: white;
			}
			"""
		)

	def handle_save(self) -> None:
		password = self.password_input.text()
		if not password or len(password) < 6:
			QMessageBox.warning(self, "Validation Error", "Password must be at least 6 characters.")
			return
		self.new_password = password
		self.accept()


class DarkTableDelegate(QStyledItemDelegate):
	def createEditor(self, parent, option, index):
		editor = QLineEdit(parent)
		editor.setStyleSheet("""
			QLineEdit {
				background-color: #0f172a;
				color: #f8fafc;
				border: 2px solid #3b82f6;
				border-radius: 4px;
				padding: 2px 6px;
			}
		""")
		return editor


class ProductDialog(QDialog):
	def __init__(self, database: AuthenticationDatabase, product: Optional[Tuple] = None, parent: Optional[QWidget] = None) -> None:
		super().__init__(parent)
		self.database = database
		self.product = product  # If editing, this is (id, sku, name, category, quantity, unit_price, min_stock, supplier)
		self.is_edit = product is not None

		self.setWindowTitle("Edit Product" if self.is_edit else "Add New Product")
		self.setMinimumSize(400, 450)
		self._build_ui()
		
		# Apply focus glow filter
		self.focus_glow_filter = FocusGlowFilter(self)
		apply_focus_glow(self, self.focus_glow_filter)

		self._load_categories()

		if self.is_edit:
			self._populate_fields()

	def _build_ui(self) -> None:
		layout = QVBoxLayout(self)
		layout.setContentsMargins(24, 24, 24, 24)
		layout.setSpacing(16)

		title = QLabel("Product Information")
		title.setStyleSheet("font-size: 18px; font-weight: 700; color: #ffffff;")
		layout.addWidget(title)

		form = QFormLayout()
		form.setVerticalSpacing(12)
		form.setHorizontalSpacing(14)

		self.sku_input = QLineEdit()
		self.sku_input.setPlaceholderText("e.g. PROD-1001")
		if self.is_edit:
			self.sku_input.setEnabled(False)  # SKU cannot be changed once created

		self.name_input = QLineEdit()
		self.name_input.setPlaceholderText("e.g. Wireless Mouse")

		self.category_input = QComboBox()
		self.category_input.setEditable(True)
		self.category_input.setPlaceholderText("Select or type category")

		self.quantity_input = QSpinBox()
		self.quantity_input.setRange(0, 1000000)
		self.quantity_input.setValue(0)
		if self.is_edit:
			self.quantity_input.setEnabled(False)  # Quantity edits must go through Adjust Stock

		self.price_input = QDoubleSpinBox()
		self.price_input.setRange(0.00, 1000000.00)
		self.price_input.setDecimals(2)
		self.price_input.setPrefix("$")
		self.price_input.setValue(0.00)

		self.min_stock_input = QSpinBox()
		self.min_stock_input.setRange(0, 10000)
		self.min_stock_input.setValue(5)

		self.supplier_input = QLineEdit()
		self.supplier_input.setPlaceholderText("e.g. Logitech Inc.")

		form.addRow("SKU Code:", self.sku_input)
		form.addRow("Product Name:", self.name_input)
		form.addRow("Category:", self.category_input)
		form.addRow("Initial Quantity:", self.quantity_input)
		form.addRow("Unit Price ($):", self.price_input)
		form.addRow("Min Stock Warning:", self.min_stock_input)
		form.addRow("Supplier Details:", self.supplier_input)

		layout.addLayout(form)

		button_row = QHBoxLayout()
		self.cancel_button = QPushButton("Cancel")
		self.cancel_button.setObjectName("dialogCancelBtn")
		self.cancel_button.clicked.connect(self.reject)

		self.save_button = QPushButton("Save Product")
		self.save_button.setObjectName("dialogSaveBtn")
		self.save_button.clicked.connect(self.handle_save)

		button_row.addStretch()
		button_row.addWidget(self.cancel_button)
		button_row.addWidget(self.save_button)
		layout.addLayout(button_row)

		self.setStyleSheet(
			"""
			QDialog {
				background-color: #111827;
			}
			QLabel {
				color: #cbd5e1;
				font-size: 13px;
				font-weight: 500;
			}
			QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox {
				background-color: #1e293b;
				color: #f8fafc;
				border: 1px solid #334155;
				border-radius: 6px;
				padding: 6px 10px;
				font-size: 13px;
				min-height: 20px;
			}
			QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus {
				border: 1px solid #3b82f6;
			}
			QPushButton {
				border: none;
				border-radius: 6px;
				padding: 10px 18px;
				font-weight: 600;
				font-size: 13px;
			}
			#dialogSaveBtn {
				background-color: #3b82f6;
				color: white;
			}
			#dialogSaveBtn:hover {
				background-color: #2563eb;
			}
			#dialogCancelBtn {
				background-color: #1e293b;
				color: #cbd5e1;
				border: 1px solid #334155;
			}
			#dialogCancelBtn:hover {
				background-color: #334155;
				color: white;
			}
			"""
		)

	def _load_categories(self) -> None:
		try:
			connection = self.database._connect()
			cursor = connection.cursor()
			query = self.database._format_query("SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category != ''")
			cursor.execute(query)
			rows = cursor.fetchall()
			categories = [row[0] for row in rows]
			self.category_input.addItems(categories)
			cursor.close()
			connection.close()
		except Exception:
			pass  # Non-blocking if it fails on empty database

	def _populate_fields(self) -> None:
		if not self.product:
			return
		# product = (id, sku, name, category, quantity, unit_price, min_stock, supplier)
		self.sku_input.setText(self.product[1])
		self.name_input.setText(self.product[2])
		self.category_input.setEditText(self.product[3] or "")
		self.quantity_input.setValue(self.product[4] or 0)
		self.price_input.setValue(float(self.product[5] or 0.0))
		self.min_stock_input.setValue(self.product[6] or 0)
		self.supplier_input.setText(self.product[7] or "")

	def handle_save(self) -> None:
		sku = self.sku_input.text().strip().upper()
		name = self.name_input.text().strip()
		category = self.category_input.currentText().strip()
		quantity = self.quantity_input.value()
		price = self.price_input.value()
		min_stock = self.min_stock_input.value()
		supplier = self.supplier_input.text().strip()

		if not sku or not name:
			QMessageBox.warning(self, "Validation Error", "SKU Code and Product Name are required fields.")
			return

		self.save_data = {
			"sku": sku,
			"name": name,
			"category": category,
			"quantity": quantity,
			"price": price,
			"min_stock": min_stock,
			"supplier": supplier
		}
		self.accept()


class StockAdjustDialog(QDialog):
	def __init__(self, parent: QWidget, product_id: int, product_sku: str, product_name: str, current_stock: int) -> None:
		super().__init__(parent)
		self.product_id = product_id
		self.product_sku = product_sku
		self.product_name = product_name
		self.current_stock = current_stock

		self.setWindowTitle("Adjust Stock Level")
		self.setMinimumSize(400, 320)
		self._build_ui()

		# Apply focus glow filter
		self.focus_glow_filter = FocusGlowFilter(self)
		apply_focus_glow(self, self.focus_glow_filter)

	def _build_ui(self) -> None:
		layout = QVBoxLayout(self)
		layout.setContentsMargins(24, 24, 24, 24)
		layout.setSpacing(16)

		title = QLabel("Adjust Product Stock")
		title.setStyleSheet("font-size: 18px; font-weight: 700; color: #ffffff;")
		layout.addWidget(title)

		info_frame = QFrame()
		info_frame.setStyleSheet("background-color: #1e293b; border-radius: 6px; padding: 12px;")
		info_layout = QFormLayout(info_frame)
		info_layout.setSpacing(8)

		info_layout.addRow(QLabel("Product SKU:"), QLabel(self.product_sku))
		info_layout.addRow(QLabel("Product Name:"), QLabel(self.product_name))
		
		self.stock_lbl = QLabel(str(self.current_stock))
		self.stock_lbl.setStyleSheet("font-weight: 700; color: #3b82f6;")
		info_layout.addRow(QLabel("Current Stock:"), self.stock_lbl)
		layout.addWidget(info_frame)

		form = QFormLayout()
		form.setVerticalSpacing(12)
		
		self.action_combo = QComboBox()
		self.action_combo.addItems(["Add Stock (Stock In)", "Remove Stock (Stock Out)"])
		
		self.quantity_input = QSpinBox()
		self.quantity_input.setRange(1, 100000)
		self.quantity_input.setValue(10)

		self.notes_input = QLineEdit()
		self.notes_input.setPlaceholderText("e.g. Received new shipment, Damaged returned")

		form.addRow("Adjustment Action:", self.action_combo)
		form.addRow("Quantity to Adjust:", self.quantity_input)
		form.addRow("Reason / Notes:", self.notes_input)
		layout.addLayout(form)

		button_row = QHBoxLayout()
		self.cancel_button = QPushButton("Cancel")
		self.cancel_button.setObjectName("dialogCancelBtn")
		self.cancel_button.clicked.connect(self.reject)

		self.save_button = QPushButton("Confirm Adjustment")
		self.save_button.setObjectName("dialogSaveBtn")
		self.save_button.clicked.connect(self.handle_save)

		button_row.addStretch()
		button_row.addWidget(self.cancel_button)
		button_row.addWidget(self.save_button)
		layout.addLayout(button_row)

		self.setStyleSheet(
			"""
			QDialog {
				background-color: #111827;
			}
			QLabel {
				color: #cbd5e1;
				font-size: 13px;
				font-weight: 500;
			}
			QLineEdit, QComboBox, QSpinBox {
				background-color: #1e293b;
				color: #f8fafc;
				border: 1px solid #334155;
				border-radius: 6px;
				padding: 6px 10px;
				font-size: 13px;
				min-height: 20px;
			}
			QLineEdit:focus, QComboBox:focus, QSpinBox:focus {
				border: 1px solid #3b82f6;
			}
			QPushButton {
				border: none;
				border-radius: 6px;
				padding: 10px 18px;
				font-weight: 600;
				font-size: 13px;
			}
			#dialogSaveBtn {
				background-color: #3b82f6;
				color: white;
			}
			#dialogSaveBtn:hover {
				background-color: #2563eb;
			}
			#dialogCancelBtn {
				background-color: #1e293b;
				color: #cbd5e1;
				border: 1px solid #334155;
			}
			#dialogCancelBtn:hover {
				background-color: #334155;
				color: white;
			}
			"""
		)

	def handle_save(self) -> None:
		action = "STOCK_IN" if self.action_combo.currentIndex() == 0 else "STOCK_OUT"
		qty = self.quantity_input.value()
		notes = self.notes_input.text().strip()

		if action == "STOCK_OUT" and qty > self.current_stock:
			QMessageBox.warning(self, "Invalid Quantity", "Cannot remove more stock than currently available in inventory.")
			return

		self.adjustment_data = {
			"action": action,
			"qty": qty,
			"notes": notes if notes else ("Stock adjustment: " + ("In" if action == "STOCK_IN" else "Out"))
		}
		self.accept()


class InventoryWindow(QMainWindow):
	def __init__(self, database: AuthenticationDatabase, logged_in_user: str) -> None:
		super().__init__()
		self.database = database
		self.logged_in_user = logged_in_user
		self.user_role = self.database.get_user_role(logged_in_user)

		self.setWindowTitle("PySide Inventory Management")
		self.setMinimumSize(700, 520)
		self._sidebar_collapsed = False
		self.active_tab = 0

		self.current_page = 1
		self.page_limit = 25

		self.user_current_page = 1
		self.user_page_limit = 20

		self._build_ui()
		
		# Apply focus glow filter
		self.focus_glow_filter = FocusGlowFilter(self)
		apply_focus_glow(self, self.focus_glow_filter)

		# Initial Data Load
		self.refresh_all_data()
   








   
	def _build_ui(self) -> None:
		central = QWidget()
		self.setCentralWidget(central)

		main_layout = QHBoxLayout(central)
		main_layout.setContentsMargins(0, 0, 0, 0)
		main_layout.setSpacing(0)

		# ------------------ SIDEBAR ------------------
		self.sidebar = QFrame()
		sidebar = self.sidebar
		sidebar.setObjectName("sidebarFrame")
		sidebar_layout = QVBoxLayout(sidebar)
		sidebar_layout.setContentsMargins(16, 24, 16, 24)
		sidebar_layout.setSpacing(8)

		app_title = QLabel("Inventory Admin")
		app_title.setObjectName("sidebarTitle")
		sidebar_layout.addWidget(app_title)

		user_lbl = QLabel(f"Logged in as: {self.logged_in_user}")
		user_lbl.setObjectName("sidebarUser")
		sidebar_layout.addWidget(user_lbl)

		# Navigation Buttons
		self.nav_buttons: List[QPushButton] = []
		menu_items = [
			("Dashboard", 0),
			("Product Management", 1),
		]
		if self.user_role == 'admin':
			menu_items.append(("User List", 2))

		nav_tooltips = {
			0: "View overview statistics and KPIs",
			1: "Add, edit, delete, or adjust product inventory",
			2: "View list of registered users"
		}

		for label, idx in menu_items:
			btn = QPushButton(label)
			btn.setProperty("idx", idx)
			btn.setCheckable(True)
			btn.setAutoExclusive(True)
			btn.setCursor(Qt.PointingHandCursor)
			btn.setObjectName("sidebarBtn")
			btn.clicked.connect(self.handle_nav_click)
			btn.setToolTip(nav_tooltips.get(idx, ""))
			sidebar_layout.addWidget(btn)
			self.nav_buttons.append(btn)

		# Set Dashboard as active initially
		self.nav_buttons[0].setChecked(True)
		self.nav_buttons[0].setStyleSheet("background-color: #3b82f6; color: #ffffff; font-weight: 600;")

		sidebar_layout.addStretch()

		# Logout Button
		logout_btn = QPushButton("Logout")
		logout_btn.setObjectName("logoutBtn")
		logout_btn.setCursor(Qt.PointingHandCursor)
		logout_btn.clicked.connect(self.handle_logout)
		logout_btn.setToolTip("Logout and end current session")
		sidebar_layout.addWidget(logout_btn)

		main_layout.addWidget(sidebar)

		# ------------------ CONTENT AREA ------------------
		content_frame = QFrame()
		self._content_layout = QVBoxLayout(content_frame)
		content_layout = self._content_layout
		content_layout.setContentsMargins(28, 24, 28, 24)
		content_layout.setSpacing(20)

		# Content Header
		header_layout = QHBoxLayout()

		self.toggle_sidebar_btn = QPushButton("☰")
		self.toggle_sidebar_btn.setCursor(Qt.PointingHandCursor)
		self.toggle_sidebar_btn.setObjectName("toggleSidebarBtn")
		self.toggle_sidebar_btn.clicked.connect(self.toggle_sidebar)
		self.toggle_sidebar_btn.setToolTip("Toggle sidebar visibility")
		self.toggle_sidebar_btn.setStyleSheet("""
			QPushButton {
				background-color: transparent;
				color: #ffffff;
				font-size: 20px;
				border: none;
				padding: 4px;
				margin-right: 12px;
			}
			QPushButton:hover {
				color: #3b82f6;
			}
		""")
		header_layout.addWidget(self.toggle_sidebar_btn)

		self.header_title = QLabel("Dashboard Overview")
		self.header_title.setObjectName("headerTitle")
		header_layout.addWidget(self.header_title)
		header_layout.addStretch()
		content_layout.addLayout(header_layout)

		# Stacked Widget
		self.stacked_widget = QStackedWidget()
		self.stacked_widget.addWidget(self._build_dashboard_page())
		self.stacked_widget.addWidget(self._build_inventory_page())
		self.stacked_widget.addWidget(self._build_users_page())

		content_layout.addWidget(self.stacked_widget)
		main_layout.addWidget(content_frame)

		# ------------------ STYLING ------------------
		self.setStyleSheet(
			"""
			QMainWindow {
				background-color: #0f172a;
			}
			QWidget {
				font-family: "Segoe UI", -apple-system, Roboto, Helvetica, sans-serif;
			}
			#sidebarFrame {
				background-color: #0b0f19;
				border-right: 1px solid #1e293b;
				min-width: 220px;
				max-width: 220px;
			}
			#sidebarFrame[collapsed="true"] {
				min-width: 0px;
				max-width: 0px;
				border-right: none;
			}
			#sidebarTitle {
				font-size: 20px;
				font-weight: 700;
				color: #ffffff;
				padding: 10px 8px 2px 8px;
			}
			#sidebarUser {
				font-size: 12px;
				color: #64748b;
				padding: 0px 8px 16px 8px;
				margin-bottom: 20px;
				border-bottom: 1px solid #1e293b;
			}
			#sidebarBtn {
				background-color: transparent;
				color: #94a3b8;
				border: none;
				border-radius: 6px;
				padding: 12px 14px;
				text-align: left;
				font-size: 13px;
				font-weight: 500;
			}
			#sidebarBtn:hover {
				background-color: #1e293b;
				color: #f8fafc;
			}
			#sidebarBtn:checked {
				background-color: #3b82f6;
				color: #ffffff;
				font-weight: 600;
			}
			#logoutBtn {
				background-color: #1e293b;
				color: #f43f5e;
				border: none;
				border-radius: 6px;
				padding: 10px 14px;
				font-weight: 600;
				font-size: 13px;
			}
			#logoutBtn:hover {
				background-color: #e11d48;
				color: #ffffff;
			}
			#headerTitle {
				font-size: 24px;
				font-weight: 700;
				color: #ffffff;
			}
			QLineEdit, QComboBox {
				background-color: #1e293b;
				color: #f8fafc;
				border: 1px solid #334155;
				border-radius: 6px;
				padding: 6px 12px;
				font-size: 13px;
				min-height: 20px;
			}
			QLineEdit:focus, QComboBox:focus {
				border: 1px solid #3b82f6;
				background-color: #1e293b;
			}
			QComboBox::drop-down {
				subcontrol-origin: padding;
				subcontrol-position: right center;
				width: 24px;
				border-left: 1px solid #334155;
				border-radius: 0 6px 6px 0;
				background-color: #1e293b;
			}
			QComboBox::down-arrow {
				width: 10px;
				height: 10px;
				border-left: 2px solid #94a3b8;
				border-bottom: 2px solid #94a3b8;
				transform: rotate(-45deg);
				margin-top: -4px;
			}
			QComboBox QAbstractItemView {
				background-color: #0f172a;
				color: #f8fafc;
				border: 1px solid #334155;
				border-radius: 6px;
				padding: 4px;
				outline: none;
				selection-background-color: #1e3a5f;
				selection-color: #60a5fa;
			}
			QComboBox QAbstractItemView::item {
				padding: 6px 10px;
				min-height: 26px;
				border-radius: 4px;
			}
			QComboBox QAbstractItemView::item:hover {
				background-color: #1e293b;
				color: #ffffff;
			}
			QScrollBar:vertical {
				background: #0f172a;
				width: 8px;
				border-radius: 4px;
			}
			QScrollBar::handle:vertical {
				background: #334155;
				border-radius: 4px;
				min-height: 20px;
			}
			QScrollBar::handle:vertical:hover {
				background: #475569;
			}
			QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
				height: 0;
			}
			QScrollBar:horizontal {
				background: #0f172a;
				height: 8px;
				border-radius: 4px;
			}
			QScrollBar::handle:horizontal {
				background: #334155;
				border-radius: 4px;
				min-width: 20px;
			}
			QScrollBar::handle:horizontal:hover {
				background: #475569;
			}
			QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
				width: 0;
			}
			QTableWidget {
				background-color: #1e293b;
				color: #e2e8f0;
				border: 1px solid #334155;
				border-radius: 8px;
				gridline-color: #334155;
				outline: none;
			}
			QTableWidget::item {
				padding: 8px;
				border-bottom: 1px solid #334155;
				outline: none;
			}
			QTableWidget::item:selected {
				background-color: #334155;
				color: #60a5fa;
				outline: none;
			}
			QHeaderView::section {
				background-color: #334155;
				color: #ffffff;
				padding: 10px;
				border: none;
				font-weight: 600;
				font-size: 13px;
			}
			QPushButton {
				outline: none;
			}
			#addBtn {
				background-color: #3b82f6;
				color: white;
				border: none;
				border-radius: 6px;
				padding: 8px 16px;
				font-weight: 600;
			}
			#addBtn:hover {
				background-color: #2563eb;
			}
			#addBtn:pressed {
				background-color: #1d4ed8;
			}
			#adjustBtn {
				background-color: #10b981;
				color: white;
				border: none;
				border-radius: 6px;
				padding: 8px 16px;
				font-weight: 600;
			}
			#adjustBtn:hover {
				background-color: #059669;
			}
			#adjustBtn:pressed {
				background-color: #047857;
			}
			#editBtn {
				background-color: #1e293b;
				color: #cbd5e1;
				border: 1px solid #334155;
				border-radius: 6px;
				padding: 8px 16px;
				font-weight: 600;
			}
			#editBtn:hover {
				background-color: #334155;
				color: white;
			}
			#editBtn:pressed {
				background-color: #475569;
			}
			#deleteBtn {
				background-color: #f43f5e;
				color: white;
				border: none;
				border-radius: 6px;
				padding: 8px 16px;
				font-weight: 600;
			}
			#deleteBtn:hover {
				background-color: #e11d48;
			}
			#deleteBtn:pressed {
				background-color: #be123c;
			}
			#exportBtn {
				background-color: #8b5cf6;
				color: white;
				border: none;
				border-radius: 6px;
				padding: 8px 16px;
				font-weight: 600;
			}
			#exportBtn:hover {
				background-color: #7c3aed;
			}
			#exportBtn:pressed {
				background-color: #6d28d9;
			}
			#importBtn {
				background-color: #06b6d4;
				color: white;
				border: none;
				border-radius: 6px;
				padding: 8px 16px;
				font-weight: 600;
			}
			#importBtn:hover {
				background-color: #0891b2;
			}
			#importBtn:pressed {
				background-color: #0e7490;
			}
			"""
		)

	# ------------------ PAGES CONSTRUCTION ------------------
	def _build_dashboard_page(self) -> QWidget:
		widget = QWidget()
		layout = QVBoxLayout(widget)
		layout.setContentsMargins(0, 0, 0, 0)
		layout.setSpacing(24)

		# Welcome message with logged-in username
		self.welcome_label = QLabel(f"Welcome back, {self.logged_in_user}!")
		self.welcome_label.setStyleSheet("font-size: 20px; font-weight: 700; color: #ffffff; padding-bottom: 8px;")
		layout.addWidget(self.welcome_label)

		# KPI Cards Grid (showing only Total Users and Total Products)
		kpi_layout = QHBoxLayout()
		kpi_layout.setSpacing(20)

		self.kpi_total_users = self._create_kpi_card("Total Users", "0 Users", "#f59e0b")
		self.kpi_total_products = self._create_kpi_card("Total Products", "0 Items", "#3b82f6")

		kpi_layout.addWidget(self.kpi_total_users, 1)
		kpi_layout.addWidget(self.kpi_total_products, 1)
		layout.addLayout(kpi_layout)
		layout.addStretch()

		return widget

	def _create_kpi_card(self, label: str, val: str, accent_color: str) -> QFrame:
		frame = QFrame()
		frame.setObjectName("kpiCard")
		frame.setStyleSheet(
			f"""
			#kpiCard {{
				background-color: #1e293b;
				border: 1px solid #334155;
				border-left: 5px solid {accent_color};
				border-radius: 8px;
				padding: 20px;
				min-width: 160px;
			}}
			"""
		)
	
		layout = QVBoxLayout(frame)
		layout.setSpacing(6)

		lbl = QLabel(label)
		lbl.setStyleSheet("color: #94a3b8; font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;")
		val_lbl = QLabel(val)
		val_lbl.setStyleSheet(f"font-size: 32px; font-weight: 800; color: #ffffff;")

		layout.addWidget(lbl)
		layout.addWidget(val_lbl)

		# Store ref to value label to update dynamically later
		frame.value_label = val_lbl
		return frame

	def _build_inventory_page(self) -> QWidget:
		widget = QWidget()
		layout = QVBoxLayout(widget)
		layout.setContentsMargins(0, 0, 0, 0)
		layout.setSpacing(14)

		# Toolbar — two rows: filters on top, actions below
		toolbar_container = QVBoxLayout()
		toolbar_container.setSpacing(8)

		# Row 1: Search + filters
		filter_row = QHBoxLayout()
		filter_row.setSpacing(8)

		self.search_input = QLineEdit()
		self.search_input.setPlaceholderText("Search by SKU or Product Name...")
		self.search_input.setMinimumWidth(180)
		self.search_input.textChanged.connect(self.reset_and_load_items)

		# Category Filter Combobox
		self.category_filter = QComboBox()
		self.category_filter.setMinimumWidth(130)
		self.category_filter.setMaximumWidth(200)
		self.category_filter.currentIndexChanged.connect(self.reset_and_load_items)

		# Sort By Combobox
		self.sort_by_combobox = QComboBox()
		self.sort_by_combobox.setMinimumWidth(140)
		self.sort_by_combobox.setMaximumWidth(220)
		self.sort_by_combobox.addItems([
			"Sort: SKU (A-Z)",
			"Sort: SKU (Z-A)",
			"Sort: Name (A-Z)",
			"Sort: Name (Z-A)",
			"Sort: Quantity (Low→High)",
			"Sort: Quantity (High→Low)",
			"Sort: Price (Low→High)",
			"Sort: Price (High→Low)"
		])
		self.sort_by_combobox.currentIndexChanged.connect(self.reset_and_load_items)

		filter_row.addWidget(self.search_input, 1)
		filter_row.addWidget(self.category_filter)
		filter_row.addWidget(self.sort_by_combobox)

		# Row 2: Action buttons
		action_row = QHBoxLayout()
		action_row.setSpacing(8)

		add_btn = QPushButton("＋ Add")
		add_btn.setObjectName("addBtn")
		add_btn.clicked.connect(self.handle_add_product)
		add_btn.setCursor(Qt.PointingHandCursor)
		add_btn.setToolTip("Add a new product to the inventory database")

		adjust_btn = QPushButton("⇅ Stock")
		adjust_btn.setObjectName("adjustBtn")
		adjust_btn.clicked.connect(self.handle_adjust_stock)
		adjust_btn.setCursor(Qt.PointingHandCursor)
		adjust_btn.setToolTip("Inward (Stock In) or Outward (Stock Out) quantity adjustments")

		edit_btn = QPushButton("✎ Edit")
		edit_btn.setObjectName("editBtn")
		edit_btn.clicked.connect(self.handle_edit_product)
		edit_btn.setCursor(Qt.PointingHandCursor)
		edit_btn.setToolTip("Edit details of the selected product")

		delete_btn = QPushButton("✕ Delete")
		delete_btn.setObjectName("deleteBtn")
		delete_btn.clicked.connect(self.handle_delete_product)
		delete_btn.setCursor(Qt.PointingHandCursor)
		delete_btn.setToolTip("Permanently delete the selected product from inventory")

		export_btn = QPushButton("↓ Excel")
		export_btn.setObjectName("exportBtn")
		export_btn.clicked.connect(self.handle_export_products)
		export_btn.setCursor(Qt.PointingHandCursor)
		export_btn.setToolTip("Export the product list to an Excel sheet")

		import_btn = QPushButton("↑ Import")
		import_btn.setObjectName("importBtn")
		import_btn.clicked.connect(self.handle_import_products)
		import_btn.setCursor(Qt.PointingHandCursor)
		import_btn.setToolTip("Import products from an Excel sheet")

		action_row.addWidget(add_btn)
		action_row.addWidget(adjust_btn)
		action_row.addWidget(edit_btn)
		action_row.addWidget(delete_btn)
		action_row.addWidget(export_btn)
		action_row.addWidget(import_btn)
		action_row.addStretch()

		toolbar_container.addLayout(filter_row)
		toolbar_container.addLayout(action_row)
		layout.addLayout(toolbar_container)

		# Table
		self.inventory_table = QTableWidget()
		self.inventory_table.setItemDelegate(DarkTableDelegate())
		self.inventory_table.setColumnCount(9)
		self.inventory_table.setHorizontalHeaderLabels(["Sr. No.", "ID", "SKU Code", "Product Name", "Category", "Quantity", "Price", "Min Alert", "Supplier"])
		self.inventory_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
		self.inventory_table.setSelectionBehavior(QAbstractItemView.SelectRows)
		self.inventory_table.setSelectionMode(QAbstractItemView.SingleSelection)
		self.inventory_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
		self.inventory_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Sr. No.
		self.inventory_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)  # ID
		self.inventory_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)  # SKU
		self.inventory_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Quantity
		self.inventory_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Price
		self.inventory_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)  # Min Alert
		self.inventory_table.doubleClicked.connect(self.handle_edit_product)
		# Hide row-number sidebar; enable horizontal scroll for narrow windows
		self.inventory_table.verticalHeader().setVisible(False)
		self.inventory_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
		self.inventory_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
		layout.addWidget(self.inventory_table)

		# Bottom Footer Layout (Row limit selector on left, Pagination on right)
		footer_layout = QHBoxLayout()
		
		row_limit_label = QLabel("Rows per page:")
		row_limit_label.setStyleSheet("color: #64748b; font-size: 13px; font-weight: 500;")
		self.row_limit_combo = QComboBox()
		self.row_limit_combo.addItems(["10", "25", "50", "100"])
		self.row_limit_combo.setCurrentText(str(self.page_limit))
		self.row_limit_combo.setFixedWidth(70)
		self.row_limit_combo.currentIndexChanged.connect(self.handle_row_limit_change)
		
		footer_layout.addWidget(row_limit_label)
		footer_layout.addWidget(self.row_limit_combo)
		
		# Pagination Controls — numbered page links
		self.inv_pagination_layout = QHBoxLayout()
		self.inv_pagination_layout.setSpacing(4)
		self.inv_pagination_layout.addStretch()
		
		footer_layout.addLayout(self.inv_pagination_layout, 1)
		layout.addLayout(footer_layout)

		return widget

	def _build_users_page(self) -> QWidget:
		widget = QWidget()
		layout = QVBoxLayout(widget)
		layout.setContentsMargins(0, 0, 0, 0)
		layout.setSpacing(14)

		# Toolbar
		toolbar_container = QVBoxLayout()
		toolbar_container.setSpacing(8)

		# Search
		filter_row = QHBoxLayout()
		filter_row.setSpacing(8)
		self.user_search_input = QLineEdit()
		self.user_search_input.setPlaceholderText("Search users by username or full name...")
		self.user_search_input.textChanged.connect(self.handle_user_search_change)
		filter_row.addWidget(self.user_search_input, 1)

		# Actions
		action_row = QHBoxLayout()
		action_row.setSpacing(8)

		add_user_btn = QPushButton("＋ Add User")
		add_user_btn.setObjectName("addBtn")
		add_user_btn.setCursor(Qt.PointingHandCursor)
		add_user_btn.clicked.connect(self.handle_add_user)

		delete_user_btn = QPushButton("✕ Delete User")
		delete_user_btn.setObjectName("deleteBtn")
		delete_user_btn.setCursor(Qt.PointingHandCursor)
		delete_user_btn.clicked.connect(self.handle_delete_user)

		reset_password_btn = QPushButton("🔑 Reset Password")
		reset_password_btn.setObjectName("editBtn")
		reset_password_btn.setCursor(Qt.PointingHandCursor)
		reset_password_btn.clicked.connect(self.handle_reset_password)

		toggle_admin_btn = QPushButton("👑 Toggle Admin")
		toggle_admin_btn.setObjectName("exportBtn")
		toggle_admin_btn.setCursor(Qt.PointingHandCursor)
		toggle_admin_btn.clicked.connect(self.handle_toggle_admin)

		action_row.addWidget(add_user_btn)
		action_row.addWidget(delete_user_btn)
		action_row.addWidget(reset_password_btn)
		action_row.addWidget(toggle_admin_btn)
		action_row.addStretch()

		toolbar_container.addLayout(filter_row)
		toolbar_container.addLayout(action_row)
		layout.addLayout(toolbar_container)

		# Table
		self.users_table = QTableWidget()
		self.users_table.setItemDelegate(DarkTableDelegate())
		self.users_table.setColumnCount(6)
		self.users_table.setHorizontalHeaderLabels(["ID", "Full Name", "Username", "Mobile No", "Role", "Date Registered"])
		self.users_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
		self.users_table.setSelectionBehavior(QAbstractItemView.SelectRows)
		self.users_table.setSelectionMode(QAbstractItemView.SingleSelection)
		self.users_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
		self.users_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents) # ID
		self.users_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents) # Role
		layout.addWidget(self.users_table)

		# Bottom Footer Layout (Row limit selector on left, Pagination on right)
		usr_footer_layout = QHBoxLayout()
		
		usr_limit_label = QLabel("Rows per page:")
		usr_limit_label.setStyleSheet("color: #64748b; font-size: 13px; font-weight: 500;")
		self.usr_limit_combo = QComboBox()
		self.usr_limit_combo.addItems(["10", "20", "50", "100"])
		self.usr_limit_combo.setCurrentText(str(self.user_page_limit))
		self.usr_limit_combo.setFixedWidth(70)
		self.usr_limit_combo.currentIndexChanged.connect(self.handle_usr_row_limit_change)
		
		usr_footer_layout.addWidget(usr_limit_label)
		usr_footer_layout.addWidget(self.usr_limit_combo)
		
		# Pagination Controls
		self.usr_pagination_layout = QHBoxLayout()
		self.usr_pagination_layout.setSpacing(4)
		self.usr_pagination_layout.addStretch()
		
		usr_footer_layout.addLayout(self.usr_pagination_layout, 1)
		layout.addLayout(usr_footer_layout)

		return widget

	# ------------------ INTERACTIVE HANDLERS ------------------
	# ------------------------------------------------------------------
	# Responsive layout helpers
	# ------------------------------------------------------------------

	def resizeEvent(self, event) -> None:  # noqa: N802
		super().resizeEvent(event)
		w = self.width()
		# Auto-collapse sidebar below 900 px, expand above
		if w < 900 and not self._sidebar_collapsed:
			self._set_sidebar_collapsed(True)
		elif w >= 900 and self._sidebar_collapsed:
			self._set_sidebar_collapsed(False)
		# Adjust content padding
		pad = 14 if w < 900 else 28
		self._content_layout.setContentsMargins(pad, 20, pad, 20)
		# Scale header font
		fs = 18 if w < 900 else 24
		self.header_title.setStyleSheet(
			f"font-size: {fs}px; font-weight: 700; color: #ffffff;"
		)

	def _set_sidebar_collapsed(self, collapsed: bool) -> None:
		self._sidebar_collapsed = collapsed
		self.sidebar.setProperty("collapsed", "true" if collapsed else "false")
		self.sidebar.style().unpolish(self.sidebar)
		self.sidebar.style().polish(self.sidebar)
		self.sidebar.setVisible(not collapsed)

	def toggle_sidebar(self) -> None:
		self._set_sidebar_collapsed(not self._sidebar_collapsed)

	def handle_nav_click(self) -> None:
		sender = self.sender()
		if not sender:
			return
		idx = sender.property("idx")
		self.switch_page(idx)

	def switch_page(self, idx: int) -> None:
		self.active_tab = idx
		self.stacked_widget.setCurrentIndex(idx)
		
		# Reset stylesheet on all nav buttons to clear active state
		for i, btn in enumerate(self.nav_buttons):
			if i == idx:
				btn.setChecked(True)
				btn.setStyleSheet("background-color: #3b82f6; color: #ffffff; font-weight: 600;")
			else:
				btn.setChecked(False)
				btn.setStyleSheet("")

		# Update header title based on active index
		headers = ["Dashboard Overview", "Product Management", "Registered System Users"]
		if idx < len(headers):
			self.header_title.setText(headers[idx])

		# Trigger view refresh
		if idx == 0:
			self.load_dashboard_kpis()
		elif idx == 1:
			self.load_inventory_items()
			self.refresh_categories_combobox()
		elif idx == 2:
			self.load_users()

	def refresh_all_data(self) -> None:
		self.load_dashboard_kpis()
		self.load_inventory_items()
		self.refresh_categories_combobox()
		self.load_users()

	# ------------------ REFRESH / DATA LOADING LOGIC ------------------
	def load_dashboard_kpis(self) -> None:
		try:
			connection = self.database._connect()
			cursor = connection.cursor()

			# 0. Total Users Count
			query_users = self.database._format_query("SELECT COUNT(*) FROM users")
			cursor.execute(query_users)
			total_users = cursor.fetchone()[0]

			# 1. Total Products Count
			query_total = self.database._format_query("SELECT COUNT(*) FROM products")
			cursor.execute(query_total)
			total_prods = cursor.fetchone()[0]

			# Set KPI card values
			self.kpi_total_users.value_label.setText(f"{total_users} Users")
			self.kpi_total_products.value_label.setText(f"{total_prods} Items")

			cursor.close()
			connection.close()
		except Exception as exc:
			print(f"Error loading KPI dashboard data: {exc}")

	def reset_and_load_items(self) -> None:
		self.current_page = 1
		self.load_inventory_items()

	def handle_prev_page(self) -> None:
		if self.current_page > 1:
			self.current_page -= 1
			self.load_inventory_items()

	def handle_next_page(self) -> None:
		self.current_page += 1
		self.load_inventory_items()

	def handle_user_prev_page(self) -> None:
		if self.user_current_page > 1:
			self.user_current_page -= 1
			self.load_users()

	def handle_user_next_page(self) -> None:
		self.user_current_page += 1
		self.load_users()

	# ------------------------------------------------------------------
	# Shared pagination-link builder
	# ------------------------------------------------------------------

	_PAGE_BTN_BASE = """
		QPushButton {
			background-color: #1e293b;
			color: #94a3b8;
			border: 1px solid #334155;
			border-radius: 6px;
			padding: 5px 10px;
			font-size: 13px;
			font-weight: 600;
			min-width: 34px;
		}
		QPushButton:hover {
			background-color: #334155;
			color: #f8fafc;
		}
		QPushButton:disabled {
			background-color: #0f172a;
			color: #475569;
			border: 1px solid #1e293b;
		}
	"""

	_PAGE_BTN_ACTIVE = """
		QPushButton {
			background-color: #3b82f6;
			color: #ffffff;
			border: 1px solid #3b82f6;
			border-radius: 6px;
			padding: 5px 10px;
			font-size: 13px;
			font-weight: 700;
			min-width: 34px;
		}
	"""

	_PAGE_BTN_NAV = """
		QPushButton {
			background-color: #1e293b;
			color: #60a5fa;
			border: 1px solid #334155;
			border-radius: 6px;
			padding: 5px 12px;
			font-size: 14px;
			font-weight: 700;
			min-width: 34px;
		}
		QPushButton:hover {
			background-color: #3b82f6;
			color: #ffffff;
			border: 1px solid #3b82f6;
		}
		QPushButton:disabled {
			background-color: #0f172a;
			color: #334155;
			border: 1px solid #1e293b;
		}
	"""

	def _build_page_links(
		self,
		layout: QHBoxLayout,
		current_page: int,
		total_pages: int,
		goto_callback,
	) -> None:
		"""Clear *layout* and rebuild numbered page-link buttons.

		Shows at most 9 buttons with '…' ellipsis gaps when there are many pages.
		Layout already has leading and trailing stretches inserted at build time.
		"""
		# Remove all widgets (leave the single leading stretch at index 0)
		while layout.count() > 1:
			item = layout.takeAt(1)
			if item and item.widget():
				item.widget().deleteLater()

		insert_at = 1  # insert after the leading stretch

		def _add(widget):
			nonlocal insert_at
			layout.insertWidget(insert_at, widget)
			insert_at += 1

		def _make_page_btn(page_num: int) -> QPushButton:
			btn = QPushButton(str(page_num))
			btn.setCursor(Qt.PointingHandCursor)
			if page_num == current_page:
				btn.setStyleSheet(self._PAGE_BTN_ACTIVE)
				btn.setEnabled(False)
			else:
				btn.setStyleSheet(self._PAGE_BTN_BASE)
				btn.clicked.connect(lambda checked, p=page_num: goto_callback(p))
			return btn

		def _make_ellipsis() -> QPushButton:
			btn = QPushButton("…")
			btn.setEnabled(False)
			btn.setStyleSheet(self._PAGE_BTN_BASE)
			return btn

		# ‹ Prev arrow
		prev_btn = QPushButton("‹")
		prev_btn.setCursor(Qt.PointingHandCursor)
		prev_btn.setStyleSheet(self._PAGE_BTN_NAV)
		prev_btn.setEnabled(current_page > 1)
		if current_page > 1:
			prev_btn.clicked.connect(lambda: goto_callback(current_page - 1))
		_add(prev_btn)

		# Determine which page numbers to show
		if total_pages <= 9:
			page_numbers = list(range(1, total_pages + 1))
		else:
			# Always show first, last, and a window around current
			window = set(range(max(1, current_page - 2), min(total_pages, current_page + 2) + 1))
			page_numbers = sorted({1, total_pages} | window)

		prev_num = None
		for p in page_numbers:
			if prev_num is not None and p - prev_num > 1:
				_add(_make_ellipsis())
			_add(_make_page_btn(p))
			prev_num = p

		# › Next arrow
		next_btn = QPushButton("›")
		next_btn.setCursor(Qt.PointingHandCursor)
		next_btn.setStyleSheet(self._PAGE_BTN_NAV)
		next_btn.setEnabled(current_page < total_pages)
		if current_page < total_pages:
			next_btn.clicked.connect(lambda: goto_callback(current_page + 1))
		_add(next_btn)

	# ------------------------------------------------------------------
	# Loading-screen helpers
	# ------------------------------------------------------------------

	def _show_inventory_loading(self) -> None:
		"""Show a loading overlay over the inventory table area."""
		if not hasattr(self, '_inv_loading') or self._inv_loading is None:
			self._inv_loading = LoadingScreen(self, "Loading products…")
			self._inv_loading.resize(300, 230)
			self._inv_loading.show()

	def _hide_inventory_loading(self) -> None:
		if hasattr(self, '_inv_loading') and self._inv_loading is not None:
			self._inv_loading.finish()
			self._inv_loading = None

	def _show_users_loading(self) -> None:
		"""Show a loading overlay over the users table area."""
		if not hasattr(self, '_usr_loading') or self._usr_loading is None:
			self._usr_loading = LoadingScreen(self, "Loading users…")
			self._usr_loading.resize(300, 230)
			self._usr_loading.show()

	def _hide_users_loading(self) -> None:
		if hasattr(self, '_usr_loading') and self._usr_loading is not None:
			self._usr_loading.finish()
			self._usr_loading = None

	def load_inventory_items(self) -> None:
		"""Kick off an async inventory fetch, showing a loading overlay."""
		search_text = self.search_input.text().strip()
		selected_category = (
			self.category_filter.currentText()
			if hasattr(self, 'category_filter') else "All Categories"
		)
		sort_idx = self.sort_by_combobox.currentIndex() if hasattr(self, 'sort_by_combobox') else 0

		self._show_inventory_loading()

		worker = _InventoryLoadWorker(
			self.database,
			self.current_page,
			self.page_limit,
			search_text,
			selected_category,
			sort_idx,
		)
		worker.finished.connect(self._on_inventory_loaded)
		worker.error.connect(self._on_inventory_error)
		# Keep a reference so the thread isn't garbage-collected
		self._inv_worker = worker
		worker.start()

	def _on_inventory_loaded(self, rows: list, page: int, total_pages: int) -> None:
		"""Populate the inventory table once the worker finishes."""
		self.current_page = page
		self._inv_total_pages = total_pages

		if hasattr(self, 'inv_pagination_layout'):
			self._build_page_links(
				self.inv_pagination_layout,
				page,
				total_pages,
				self._goto_inventory_page,
			)

		self.inventory_table.setRowCount(len(rows))
		start_sr = (page - 1) * self.page_limit + 1
		for r_idx, row in enumerate(rows):
			prod_id, sku, name, category, quantity, unit_price, min_stock, supplier = row

			# Sr. No. — grey, centered
			sr_item = QTableWidgetItem(str(start_sr + r_idx))
			sr_item.setTextAlignment(Qt.AlignCenter)
			sr_item.setForeground(QColor("#64748b"))
			sr_item.setToolTip(str(start_sr + r_idx))
			self.inventory_table.setItem(r_idx, 0, sr_item)

			item_id = QTableWidgetItem(str(prod_id))
			item_id.setToolTip(str(prod_id))
			self.inventory_table.setItem(r_idx, 1, item_id)

			item_sku = QTableWidgetItem(sku)
			item_sku.setToolTip(sku)
			self.inventory_table.setItem(r_idx, 2, item_sku)

			item_name = QTableWidgetItem(name)
			item_name.setToolTip(name)
			self.inventory_table.setItem(r_idx, 3, item_name)

			item_cat = QTableWidgetItem(category or "Unassigned")
			item_cat.setToolTip(category or "Unassigned")
			self.inventory_table.setItem(r_idx, 4, item_cat)

			# Stock warning highlight
			qty_item = QTableWidgetItem(str(quantity))
			qty_item.setToolTip(str(quantity))
			if quantity <= (min_stock or 0):
				qty_item.setForeground(QColor("#fca5a5"))
				qty_item.setFont(QFont("Segoe UI", 9, QFont.Bold))
			self.inventory_table.setItem(r_idx, 5, qty_item)

			item_price = QTableWidgetItem(f"${unit_price:,.2f}")
			item_price.setToolTip(f"${unit_price:,.2f}")
			self.inventory_table.setItem(r_idx, 6, item_price)

			item_min = QTableWidgetItem(str(min_stock))
			item_min.setToolTip(str(min_stock))
			self.inventory_table.setItem(r_idx, 7, item_min)

			item_sup = QTableWidgetItem(supplier or "")
			item_sup.setToolTip(supplier or "")
			self.inventory_table.setItem(r_idx, 8, item_sup)

		self._hide_inventory_loading()

	def _on_inventory_error(self, message: str) -> None:
		self._hide_inventory_loading()
		QMessageBox.critical(self, "Query Error", f"Unable to retrieve inventory data: {message}")

	def refresh_categories_combobox(self) -> None:
		if not hasattr(self, 'category_filter'):
			return

		current_category = self.category_filter.currentText()
		
		self.category_filter.blockSignals(True)
		self.category_filter.clear()
		self.category_filter.addItem("All Categories")
		
		try:
			connection = self.database._connect()
			cursor = connection.cursor()
			
			# Check if there are any products with empty/null category to decide whether to add "Unassigned" option
			check_query = self.database._format_query(
				"SELECT COUNT(*) FROM products WHERE category IS NULL OR category = ''"
			)
			cursor.execute(check_query)
			has_unassigned = cursor.fetchone()[0] > 0
			
			query = self.database._format_query(
				"SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category != '' ORDER BY category ASC"
			)
			cursor.execute(query)
			rows = cursor.fetchall()
			categories = [row[0] for row in rows]
			
			self.category_filter.addItems(categories)
			if has_unassigned:
				self.category_filter.addItem("Unassigned")
				
			cursor.close()
			connection.close()
		except Exception as exc:
			print(f"Error loading categories: {exc}")
			
		# Restore previous selection if possible
		idx = self.category_filter.findText(current_category)
		if idx >= 0:
			self.category_filter.setCurrentIndex(idx)
		else:
			self.category_filter.setCurrentIndex(0)
			
		self.category_filter.blockSignals(False)

	def handle_export_products(self) -> None:
		"""Export all product data to an Excel (.xlsx) file."""
		try:
			connection = self.database._connect()
			cursor = connection.cursor()
			query = self.database._format_query(
				"SELECT id, sku, name, category, quantity, unit_price, min_stock, supplier FROM products ORDER BY sku ASC"
			)
			cursor.execute(query)
			rows = cursor.fetchall()
			cursor.close()
			connection.close()
		except Exception as exc:
			QMessageBox.critical(self, "Database Error", f"Unable to fetch product data for export: {exc}")
			return

		if not rows:
			QMessageBox.information(self, "No Data", "There are no products to export.")
			return

		# Ask user where to save
		file_path, _ = QFileDialog.getSaveFileName(
			self,
			"Export Products to Excel",
			"products_export.xlsx",
			"Excel Files (*.xlsx)"
		)

		if not file_path:
			return  # User cancelled

		try:
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = "Products"

			# Header styling
			header_font = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
			header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
			header_alignment = Alignment(horizontal="center", vertical="center")
			thin_border = Border(
				left=Side(style="thin", color="D1D5DB"),
				right=Side(style="thin", color="D1D5DB"),
				top=Side(style="thin", color="D1D5DB"),
				bottom=Side(style="thin", color="D1D5DB")
			)

			headers = ["ID", "SKU Code", "Product Name", "Category", "Quantity", "Unit Price", "Min Stock Alert", "Supplier"]
			for col_idx, header in enumerate(headers, 1):
				cell = ws.cell(row=1, column=col_idx, value=header)
				cell.font = header_font
				cell.fill = header_fill
				cell.alignment = header_alignment
				cell.border = thin_border

			# Data rows
			data_font = Font(name="Segoe UI", size=11)
			for row_idx, row in enumerate(rows, 2):
				prod_id, sku, name, category, quantity, unit_price, min_stock, supplier = row
				values = [prod_id, sku, name, category or "", quantity, float(unit_price or 0), min_stock, supplier or ""]
				for col_idx, value in enumerate(values, 1):
					cell = ws.cell(row=row_idx, column=col_idx, value=value)
					cell.font = data_font
					cell.border = thin_border

			# Format the price column as currency
			for row_idx in range(2, len(rows) + 2):
				ws.cell(row=row_idx, column=6).number_format = '$#,##0.00'

			# Auto-fit column widths
			for col_idx in range(1, len(headers) + 1):
				max_length = len(str(headers[col_idx - 1]))
				for row_idx in range(2, len(rows) + 2):
					cell_value = ws.cell(row=row_idx, column=col_idx).value
					if cell_value is not None:
						max_length = max(max_length, len(str(cell_value)))
				ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 4

			wb.save(file_path)
			QMessageBox.information(self, "Export Successful", f"Product data exported successfully to:\n{file_path}")
		except Exception as exc:
			QMessageBox.critical(self, "Export Error", f"Failed to export data: {exc}")

	def handle_import_products(self) -> None:
		"""Import product data from an Excel (.xlsx) file."""
		file_path, _ = QFileDialog.getOpenFileName(
			self,
			"Import Products from Excel",
			"",
			"Excel Files (*.xlsx)"
		)

		if not file_path:
			return  # User cancelled

		try:
			wb = openpyxl.load_workbook(file_path, read_only=True)
			ws = wb.active
			
			# Read rows
			rows = list(ws.iter_rows(values_only=True))
			if not rows:
				QMessageBox.warning(self, "Empty File", "The selected Excel file contains no data.")
				return
				
			# Find headers
			header_row = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
			
			# Map column indexes
			sku_idx = -1
			name_idx = -1
			category_idx = -1
			qty_idx = -1
			price_idx = -1
			min_stock_idx = -1
			supplier_idx = -1
			
			for col_idx, header in enumerate(header_row):
				if any(kw in header for kw in ["sku", "code"]):
					sku_idx = col_idx
				elif any(kw in header for kw in ["product name", "name"]):
					name_idx = col_idx
				elif "category" in header:
					category_idx = col_idx
				elif any(kw in header for kw in ["quantity", "qty", "stock"]):
					qty_idx = col_idx
				elif any(kw in header for kw in ["unit price", "price", "rate"]):
					price_idx = col_idx
				elif any(kw in header for kw in ["min alert", "min stock", "minimum"]):
					min_stock_idx = col_idx
				elif "supplier" in header:
					supplier_idx = col_idx
					
			# If SKU and Name indexes are not found, fallback to default order or report error
			if sku_idx == -1 or name_idx == -1:
				if len(header_row) >= 7:
					offset = 1 if "id" in header_row[0] or header_row[0].isdigit() else 0
					sku_idx = offset
					name_idx = offset + 1
					category_idx = offset + 2
					qty_idx = offset + 3
					price_idx = offset + 4
					min_stock_idx = offset + 5
					supplier_idx = offset + 6
				else:
					QMessageBox.warning(
						self, 
						"Invalid Format", 
						"Could not map columns. Please ensure headers like 'SKU Code' and 'Product Name' are present."
					)
					return

			# Insert into DB
			connection = self.database._connect()
			cursor = connection.cursor()
			
			success_count = 0
			duplicate_count = 0
			error_count = 0
			errors = []
			
			# Skip header row
			for r_idx, row in enumerate(rows[1:], 2):
				if not any(cell is not None for cell in row):  # Skip empty rows
					continue
					
				# Extract fields safely
				sku = str(row[sku_idx]).strip().upper() if (sku_idx >= 0 and sku_idx < len(row) and row[sku_idx] is not None) else ""
				name = str(row[name_idx]).strip() if (name_idx >= 0 and name_idx < len(row) and row[name_idx] is not None) else ""
				
				if not sku or not name:
					error_count += 1
					errors.append(f"Row {r_idx}: SKU and Name are required.")
					continue
					
				category = str(row[category_idx]).strip() if (category_idx >= 0 and category_idx < len(row) and row[category_idx] is not None) else ""
				
				# Quantity
				qty = 0
				if qty_idx >= 0 and qty_idx < len(row) and row[qty_idx] is not None:
					try:
						qty = int(row[qty_idx])
					except ValueError:
						pass
						
				# Price
				price = 0.0
				if price_idx >= 0 and price_idx < len(row) and row[price_idx] is not None:
					try:
						val_str = str(row[price_idx]).replace("$", "").replace(",", "").strip()
						price = float(val_str)
					except ValueError:
						pass
						
				# Min stock
				min_stock = 5
				if min_stock_idx >= 0 and min_stock_idx < len(row) and row[min_stock_idx] is not None:
					try:
						min_stock = int(row[min_stock_idx])
					except ValueError:
						pass
						
				supplier = str(row[supplier_idx]).strip() if (supplier_idx >= 0 and supplier_idx < len(row) and row[supplier_idx] is not None) else ""
				
				# Check if SKU exists
				try:
					check_query = self.database._format_query("SELECT id FROM products WHERE sku = %s")
					cursor.execute(check_query, (sku,))
					if cursor.fetchone():
						duplicate_count += 1
						continue
						
					# Insert Product
					insert_query = self.database._format_query(
						"INSERT INTO products (sku, name, category, quantity, unit_price, min_stock, supplier) VALUES (%s, %s, %s, %s, %s, %s, %s)"
					)
					cursor.execute(
						insert_query,
						(sku, name, category, qty, price, min_stock, supplier)
					)
					success_count += 1
				except Exception as e:
					error_count += 1
					errors.append(f"Row {r_idx} ({sku}): {str(e)}")
					
			connection.commit()
			cursor.close()
			connection.close()
			
			self.refresh_all_data()
			
			# Build report message
			report = f"Import completed successfully!\n\n"
			report += f"• Successfully imported: {success_count} products\n"
			if duplicate_count > 0:
				report += f"• Skipped (Duplicate SKU): {duplicate_count} products\n"
			if error_count > 0:
				report += f"• Errors encountered: {error_count} products\n"
				report += "\nFirst few errors:\n" + "\n".join(errors[:5])
				
			if success_count > 0:
				QMessageBox.information(self, "Import Status", report)
			else:
				QMessageBox.warning(self, "Import Status", report)
				
		except Exception as exc:
			QMessageBox.critical(self, "Import Error", f"Failed to import data: {exc}")

	def load_users(self) -> None:
		"""Kick off an async users fetch, showing a loading overlay."""
		self._show_users_loading()

		search_text = self.user_search_input.text().strip() if hasattr(self, 'user_search_input') else ""
		worker = _UsersLoadWorker(self.database, self.user_current_page, self.user_page_limit, search_text)
		worker.finished.connect(self._on_users_loaded)
		worker.error.connect(self._on_users_error)
		self._usr_worker = worker
		worker.start()

	def _on_users_loaded(self, rows: list, page: int, total_pages: int) -> None:
		"""Populate the users table once the worker finishes."""
		self.user_current_page = page
		self._usr_total_pages = total_pages

		if hasattr(self, 'usr_pagination_layout'):
			self._build_page_links(
				self.usr_pagination_layout,
				page,
				total_pages,
				self._goto_users_page,
			)

		self.users_table.setRowCount(len(rows))
		for r_idx, row in enumerate(rows):
			uid, uname, fullname, mobile, role, created_at = row
			
			item_uid = QTableWidgetItem(str(uid))
			item_uid.setToolTip(str(uid))
			self.users_table.setItem(r_idx, 0, item_uid)

			item_fullname = QTableWidgetItem(fullname)
			item_fullname.setToolTip(fullname)
			self.users_table.setItem(r_idx, 1, item_fullname)

			item_uname = QTableWidgetItem(uname)
			item_uname.setToolTip(uname)
			self.users_table.setItem(r_idx, 2, item_uname)

			item_mobile = QTableWidgetItem(mobile)
			item_mobile.setToolTip(mobile)
			self.users_table.setItem(r_idx, 3, item_mobile)

			item_role = QTableWidgetItem(role)
			item_role.setToolTip(role)
			self.users_table.setItem(r_idx, 4, item_role)

			item_created = QTableWidgetItem(str(created_at))
			item_created.setToolTip(str(created_at))
			self.users_table.setItem(r_idx, 5, item_created)

		self._hide_users_loading()

	def _on_users_error(self, message: str) -> None:
		self._hide_users_loading()
		QMessageBox.critical(self, "Users Loading Error", f"Could not retrieve registered users: {message}")

	def _goto_inventory_page(self, page: int) -> None:
		"""Jump directly to a specific inventory page."""
		self.current_page = page
		self.load_inventory_items()

	def _goto_users_page(self, page: int) -> None:
		"""Jump directly to a specific users page."""
		self.user_current_page = page
		self.load_users()

	# ------------------ CRUD IMPLEMENTATION ------------------
	def handle_add_product(self) -> None:
		dialog = ProductDialog(self.database, parent=self)
		if dialog.exec() == QDialog.Accepted:
			data = dialog.save_data
			try:
				connection = self.database._connect()
				cursor = connection.cursor()

				# Check unique SKU
				check_query = self.database._format_query("SELECT id FROM products WHERE sku = %s")
				cursor.execute(check_query, (data["sku"],))
				if cursor.fetchone():
					QMessageBox.warning(self, "Duplicate SKU", f"A product with SKU '{data['sku']}' already exists.")
					cursor.close()
					connection.close()
					return

				# Insert Product
				insert_query = self.database._format_query(
					"INSERT INTO products (sku, name, category, quantity, unit_price, min_stock, supplier) VALUES (%s, %s, %s, %s, %s, %s, %s)"
				)
				cursor.execute(
					insert_query,
					(data["sku"], data["name"], data["category"], data["quantity"], data["price"], data["min_stock"], data["supplier"])
				)
				
				connection.commit()
				cursor.close()
				connection.close()

				self.refresh_all_data()
				QMessageBox.information(self, "Product Added", "Product registered successfully.")
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to add product: {exc}")

	def handle_edit_product(self) -> None:
		selected_row = self.inventory_table.currentRow()
		if selected_row < 0:
			QMessageBox.warning(self, "Selection Required", "Please select a product from the directory list to edit.")
			return

		prod_id = int(self.inventory_table.item(selected_row, 1).text())
		
		# Fetch active product state from DB
		try:
			connection = self.database._connect()
			cursor = connection.cursor()
			query = self.database._format_query(
				"SELECT id, sku, name, category, quantity, unit_price, min_stock, supplier FROM products WHERE id = %s"
			)
			cursor.execute(query, (prod_id,))
			product = cursor.fetchone()
			cursor.close()
			connection.close()
		except Exception as exc:
			QMessageBox.critical(self, "Database Error", f"Unable to fetch product info: {exc}")
			return

		if not product:
			QMessageBox.warning(self, "Not Found", "Product details could not be found.")
			return

		dialog = ProductDialog(self.database, product=product, parent=self)
		if dialog.exec() == QDialog.Accepted:
			data = dialog.save_data
			try:
				connection = self.database._connect()
				cursor = connection.cursor()

				# Update Product
				update_query = self.database._format_query(
					"UPDATE products SET name = %s, category = %s, unit_price = %s, min_stock = %s, supplier = %s WHERE id = %s"
				)
				cursor.execute(
					update_query,
					(data["name"], data["category"], data["price"], data["min_stock"], data["supplier"], prod_id)
				)

				connection.commit()
				cursor.close()
				connection.close()

				self.refresh_all_data()
				QMessageBox.information(self, "Product Updated", "Product information updated successfully.")
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to update product: {exc}")

	def handle_adjust_stock(self) -> None:
		selected_row = self.inventory_table.currentRow()
		if selected_row < 0:
			QMessageBox.warning(self, "Selection Required", "Select a product from the list first to adjust stock.")
			return

		prod_id = int(self.inventory_table.item(selected_row, 1).text())
		sku = self.inventory_table.item(selected_row, 2).text()
		name = self.inventory_table.item(selected_row, 3).text()
		
		# We must reload product from DB to get the most updated stock quantity
		try:
			connection = self.database._connect()
			cursor = connection.cursor()
			query = self.database._format_query("SELECT quantity FROM products WHERE id = %s")
			cursor.execute(query, (prod_id,))
			current_stock = cursor.fetchone()[0]
			cursor.close()
			connection.close()
		except Exception as exc:
			QMessageBox.critical(self, "Database Error", f"Could not reload current stock: {exc}")
			return

		dialog = StockAdjustDialog(self, prod_id, sku, name, current_stock)
		if dialog.exec() == QDialog.Accepted:
			data = dialog.adjustment_data
			# data = {action: STOCK_IN/STOCK_OUT, qty: int, notes: str}

			qty_change = data["qty"] if data["action"] == "STOCK_IN" else -data["qty"]
			new_qty = current_stock + qty_change

			try:
				connection = self.database._connect()
				cursor = connection.cursor()

				# Update Quantity
				update_query = self.database._format_query("UPDATE products SET quantity = %s WHERE id = %s")
				cursor.execute(update_query, (new_qty, prod_id))

				connection.commit()
				cursor.close()
				connection.close()

				self.refresh_all_data()
				QMessageBox.information(self, "Stock Adjusted", f"Successfully recorded transaction. Stock is now {new_qty} units.")
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to update stock levels: {exc}")

	def handle_delete_product(self) -> None:
		selected_row = self.inventory_table.currentRow()
		if selected_row < 0:
			QMessageBox.warning(self, "Selection Required", "Select a product from the directory to delete.")
			return

		prod_id = int(self.inventory_table.item(selected_row, 1).text())
		sku = self.inventory_table.item(selected_row, 2).text()
		name = self.inventory_table.item(selected_row, 3).text()

		reply = QMessageBox.question(
			self,
			"Confirm Deletion",
			f"Are you sure you want to permanently delete product '{name}' (SKU: {sku})?",
			QMessageBox.Yes | QMessageBox.No
		)

		if reply == QMessageBox.Yes:
			try:
				connection = self.database._connect()
				cursor = connection.cursor()

				# Delete Product
				delete_query = self.database._format_query("DELETE FROM products WHERE id = %s")
				cursor.execute(delete_query, (prod_id,))

				connection.commit()
				cursor.close()
				connection.close()

				self.refresh_all_data()
				QMessageBox.information(self, "Product Deleted", "The product registry was removed successfully.")
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to delete product: {exc}")

	def handle_row_limit_change(self) -> None:
		try:
			new_limit = int(self.row_limit_combo.currentText())
			self.page_limit = new_limit
			self.reset_and_load_items()
		except Exception:
			pass

	def handle_usr_row_limit_change(self) -> None:
		try:
			new_limit = int(self.usr_limit_combo.currentText())
			self.user_page_limit = new_limit
			self.user_current_page = 1
			self.load_users()
		except Exception:
			pass

	def handle_user_search_change(self) -> None:
		self.user_current_page = 1
		self.load_users()

	def handle_add_user(self) -> None:
		dialog = AddUserDialog(parent=self)
		apply_focus_glow(dialog, self.focus_glow_filter)
		if dialog.exec() == QDialog.Accepted:
			data = dialog.save_data
			try:
				self.database.register_user(
					data["username"],
					data["fullname"],
					data["mobile"],
					data["password"],
					data["role"]
				)
				self.refresh_all_data()
				QMessageBox.information(self, "User Added", "User account registered successfully.")
			except ValueError as exc:
				QMessageBox.warning(self, "Duplicate Username", str(exc))
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to add user: {exc}")

	def handle_delete_user(self) -> None:
		selected_row = self.users_table.currentRow()
		if selected_row < 0:
			QMessageBox.warning(self, "Selection Required", "Select a user from the directory to delete.")
			return

		user_id = int(self.users_table.item(selected_row, 0).text())
		uname = self.users_table.item(selected_row, 2).text()
		fullname = self.users_table.item(selected_row, 1).text()

		if uname.lower() == self.logged_in_user.lower():
			QMessageBox.warning(self, "Action Denied", "You cannot delete your own logged-in account.")
			return

		reply = QMessageBox.question(
			self,
			"Confirm Deletion",
			f"Are you sure you want to permanently delete user account '{uname}' ({fullname})?",
			QMessageBox.Yes | QMessageBox.No
		)

		if reply == QMessageBox.Yes:
			try:
				self.database.delete_user_by_id(user_id)
				self.refresh_all_data()
				QMessageBox.information(self, "User Deleted", "The user account was successfully deleted.")
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to delete user: {exc}")

	def handle_reset_password(self) -> None:
		selected_row = self.users_table.currentRow()
		if selected_row < 0:
			QMessageBox.warning(self, "Selection Required", "Select a user from the directory list first to reset their password.")
			return

		user_id = int(self.users_table.item(selected_row, 0).text())
		uname = self.users_table.item(selected_row, 2).text()

		dialog = ResetPasswordDialog(uname, parent=self)
		apply_focus_glow(dialog, self.focus_glow_filter)
		if dialog.exec() == QDialog.Accepted:
			try:
				self.database.reset_user_password(user_id, dialog.new_password)
				QMessageBox.information(self, "Password Reset", f"The password for user '{uname}' was successfully reset.")
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to reset password: {exc}")

	def handle_toggle_admin(self) -> None:
		selected_row = self.users_table.currentRow()
		if selected_row < 0:
			QMessageBox.warning(self, "Selection Required", "Select a user from the directory to toggle their administrative privileges.")
			return

		user_id = int(self.users_table.item(selected_row, 0).text())
		uname = self.users_table.item(selected_row, 2).text()
		current_role = self.users_table.item(selected_row, 4).text()

		if uname.lower() == self.logged_in_user.lower():
			QMessageBox.warning(self, "Action Denied", "You cannot demote yourself. Another administrator must change your role.")
			return

		new_role = "user" if current_role == "admin" else "admin"

		reply = QMessageBox.question(
			self,
			"Change Privileges",
			f"Are you sure you want to change user '{uname}' role from '{current_role}' to '{new_role}'?",
			QMessageBox.Yes | QMessageBox.No
		)

		if reply == QMessageBox.Yes:
			try:
				self.database.update_user_role(user_id, new_role)
				self.refresh_all_data()
				QMessageBox.information(self, "Privileges Changed", f"Successfully updated user '{uname}' to role '{new_role}'.")
			except Exception as exc:
				QMessageBox.critical(self, "Database Error", f"Unable to update user privileges: {exc}")

	# ------------------ LOGOUT SESSION ------------------
	def handle_logout(self) -> None:
		reply = QMessageBox.question(
			self,
			"Confirm Logout",
			"Are you sure you want to log out of the inventory management session?",
			QMessageBox.Yes | QMessageBox.No
		)
		
		if reply == QMessageBox.Yes:
			from authentication import AuthenticationWindow
			self.auth_window = AuthenticationWindow()
			if self.isMinimized():
				self.auth_window.showMinimized()
			else:
				self.auth_window.show()
			self.close()
